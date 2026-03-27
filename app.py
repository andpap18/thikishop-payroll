import streamlit as st
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re
import io
import tempfile
import os

# --- Configuration ---
FILL_HEADER_GREY = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
FILL_ORANGE = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
FILL_LIGHT_ORANGE = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
BORDER_THIN = Side(style='thin', color="000000")
BORDER_THICK = Side(style='medium', color="000000")
BORDER_ALL_THIN = Border(left=BORDER_THIN, right=BORDER_THIN, top=BORDER_THIN, bottom=BORDER_THIN)

# Location colors (from schedule_transformer)
FILL_RENTIS = PatternFill(start_color="FCE4D6", fill_type="solid")
FILL_AIGALEO = PatternFill(start_color="E2EFDA", fill_type="solid") 
FILL_PEIRAIAS = PatternFill(start_color="DDEBF7", fill_type="solid")
FILL_PERISTERI = PatternFill(start_color="F4B084", fill_type="solid")

def clean_name(name):
    """Removes suffixes like (8ΩΡΟΣ), (4ΩΡΟΣ) and extra spaces."""
    if not name: return ""
    name = str(name).split('(')[0]
    return name.strip()

def parse_hours(time_str, employee_name=""):
    """Parses '09:00-17:00' to decimal hours (e.g., 8.0). Returns 0 if invalid or off."""
    if not time_str or not isinstance(time_str, str):
        return 0.0
    
    time_str = re.sub(r'\[.*?\]', '', time_str).strip()
    
    # Special case: "Α", "ΑΔΕΙΑ", or "ΑΡΓΙΑ" (leave/vacation/holiday) counts as 8 hours
    time_upper = time_str.upper()
    if time_upper == 'Α' or time_upper == 'A' or 'ΑΔΕΙΑ' in time_upper or 'ADEIA' in time_upper or 'ΑΡΓΙΑ' in time_upper or 'ARGIA' in time_upper:
        if employee_name.upper() == "ΗΛΙΑΣ ΚΑΨΑΛΗΣ":
            return 4.0
        return 8.0
    
    if '-' not in time_str:
        return 0.0
    
    try:
        start_str, end_str = time_str.split('-')
        start_parts = start_str.strip().split(':')
        end_parts = end_str.strip().split(':')
        
        start_h = int(start_parts[0]) + int(start_parts[1])/60
        end_h = int(end_parts[0]) + int(end_parts[1])/60
        
        diff = end_h - start_h
        if diff < 0: diff += 24
        return diff
    except:
        return 0.0

def has_work_content(val_str):
    """Check if cell contains actual work (not RR, ΡΕΠΟ, etc)"""
    if not val_str: return False
    val_str = str(val_str).strip().upper()
    if val_str in ["NONE", "", "RR", "ΡΕΠΟ", "ΑΝΑΡΡΩΤΙΚΗ"]: return False
    return True

def get_file_date_score(filename):
    """Parses filename for sorting."""
    months = {
        'ΙΑΝ': 1, 'ΦΕΒ': 2, 'ΜΑΡ': 3, 'ΑΠΡ': 4, 'ΜΑΙ': 5, 'ΙΟΥΝ': 6,
        'ΙΟΥΛ': 7, 'ΑΥΓ': 8, 'ΣΕΠ': 9, 'ΟΚΤ': 10, 'ΝΟΕ': 11, 'ΔΕΚ': 12
    }
    
    upper_name = filename.upper()
    match = re.search(r'(\d+)_([Α-Ω]+)', upper_name)
    if match:
        day = int(match.group(1))
        month_str = match.group(2)
        
        month_num = 0
        for m_name, m_val in months.items():
            if m_name in month_str:
                month_num = m_val
                break
        
        if month_num > 0:
            return month_num * 100 + day
            
    match_num = re.search(r'(\d+)', filename)
    if match_num:
        return int(match_num.group(1))
        
    return 99999

def process_payroll(uploaded_files, target_month):
    """Main payroll processing function."""
    
    # Greek Month Map for Date Parsing
    greek_months = {
        'ΙΑΝΟΥΑΡΙΟΥ': 1, 'ΦΕΒΡΟΥΑΡΙΟΥ': 2, 'ΜΑΡΤΙΟΥ': 3, 'ΑΠΡΙΛΙΟΥ': 4, 'ΜΑΙΟΥ': 5, 'ΜΑΪΟΥ': 5,
        'ΙΟΥΝΙΟΥ': 6, 'ΙΟΥΛΙΟΥ': 7, 'ΑΥΓΟΥΣΤΟΥ': 8, 'ΣΕΠΤΕΜΒΡΙΟυ': 9, 'ΟΚΤΩΒΡΙΟΥ': 10, 'ΝΟΕΜΒΡΙΟΥ': 11, 'ΔΕΚΕΜΒΡΙΟΥ': 12
    }
    
    # Month names for filename
    month_names = {
        1: 'ΙΑΝΟΥΑΡΙΟΣ', 2: 'ΦΕΒΡΟΥΑΡΙΟΣ', 3: 'ΜΑΡΤΙΟΣ', 4: 'ΑΠΡΙΛΙΟΣ', 
        5: 'ΜΑΙΟΣ', 6: 'ΙΟΥΝΙΟΣ', 7: 'ΙΟΥΛΙΟΣ', 8: 'ΑΥΓΟΥΣΤΟΣ',
        9: 'ΣΕΠΤΕΜΒΡΙΟΣ', 10: 'ΟΚΤΩΒΡΙΟΣ', 11: 'ΝΟΕΜΒΡΙΟΣ', 12: 'ΔΕΚΕΜΒΡΙΟΣ'
    }
    
    # Sort files
    file_list = [(f.name, f) for f in uploaded_files]
    file_list.sort(key=lambda x: get_file_date_score(x[0]))
    
    # Create output workbook
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "ΜΙΣΘΟΔΟΣΙΑ"
    
    current_row = 1
    monthly_stats = {}
    
    # Process each file
    for file_name, file_obj in file_list:
        # Save uploaded file to temp location
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            tmp.write(file_obj.getvalue())
            tmp_path = tmp.name
        
        try:
            wb_in = openpyxl.load_workbook(tmp_path)
            ws_in = wb_in.active
            
            # Date filtering logic
            last_data_col = 26
            include_col_map = {}
            
            col_ptr = 2
            dates_found = False
            
            for i in range(7):
                is_sunday = (i == 6)
                span = 1 if is_sunday else 4
                
                date_cell = ws_in.cell(row=2, column=col_ptr)
                date_val_raw = date_cell.value
                
                include_day = True
                if target_month and date_val_raw:
                    parsed_month = None
                    
                    if hasattr(date_val_raw, 'month'):
                        parsed_month = date_val_raw.month
                        dates_found = True
                    else:
                        date_val = str(date_val_raw).strip().upper()
                        
                        if '/' in date_val:
                            try:
                                parts = date_val.split('/')
                                if len(parts) >= 2:
                                    parsed_month = int(parts[1])
                                    dates_found = True
                            except: pass
                        
                        elif any(m in date_val for m in greek_months.keys()):
                            for m_name, m_val in greek_months.items():
                                if m_name in date_val:
                                    parsed_month = m_val
                                    dates_found = True
                                    break
                    
                    if parsed_month:
                        if parsed_month != target_month:
                            include_day = False
                
                for k in range(span):
                    include_col_map[col_ptr + k] = include_day
                
                col_ptr += span
            
            if not dates_found and target_month:
                for col_idx in include_col_map.keys():
                    include_col_map[col_idx] = True
            
            # Write Week Title
            ws_out.cell(row=current_row, column=1).value = file_name.replace("(ΕΠΙΘ).xlsx", "").replace(".xlsx", "")
            ws_out.cell(row=current_row, column=1).font = Font(bold=True, size=12)
            current_row += 1
            
            # Copy Headers (Rows 1-3)
            for r in range(1, 4):
                for c in range(1, last_data_col + 1):
                    cell_in = ws_in.cell(row=r, column=c)
                    cell_out = ws_out.cell(row=current_row + r - 1, column=c)
                    cell_out.value = cell_in.value
                    
                    if cell_in.has_style:
                        cell_out.font = Font(name=cell_in.font.name, size=cell_in.font.size, bold=True)
                        cell_out.alignment = Alignment(horizontal=cell_in.alignment.horizontal, vertical=cell_in.alignment.vertical, wrap_text=cell_in.alignment.wrap_text)
                        cell_out.border = BORDER_ALL_THIN
                        if cell_in.fill and cell_in.fill.start_color.index != '00000000':
                             cell_out.fill = PatternFill(start_color=cell_in.fill.start_color.index, fill_type='solid')
                    
                    if c in include_col_map and not include_col_map[c]:
                        cell_out.fill = PatternFill(start_color="EEEEEE", fill_type="solid")
                    
                    ws_out.column_dimensions[get_column_letter(c)].width = 16
            
            # Re-apply merges
            base_r = current_row
            col_ptr = 2
            for i in range(7):
                is_sunday = (i == 6)
                span = 1 if is_sunday else 4
                ws_out.merge_cells(start_row=base_r, start_column=col_ptr, end_row=base_r, end_column=col_ptr+span-1)
                ws_out.merge_cells(start_row=base_r+1, start_column=col_ptr, end_row=base_r+1, end_column=col_ptr+span-1)
                col_ptr += span
            
            # Add Calculation Headers
            calc_headers = ["ΗΜΕΡΕΣ ΕΡΓΑΣΙΑΣ", "ΩΡΕΣ/ΕΒΔΟ", "ΥΠΕΡΕΡΓΑΣΙΑ (h)", "ΥΠΕΡΩΡΙΕΣ(h)"]
            calc_col_start = last_data_col + 1
            
            for i, header in enumerate(calc_headers):
                c = ws_out.cell(row=current_row + 2, column=calc_col_start + i)
                c.value = header
                c.font = Font(bold=True, size=9)
                c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                c.border = BORDER_ALL_THIN
                if i == 0: c.fill = PatternFill(start_color="FFFFFF", fill_type="solid")
                elif i == 1: c.fill = PatternFill(start_color="FFFFFF", fill_type="solid")
                elif i == 2: c.fill = FILL_ORANGE
                elif i == 3: c.fill = FILL_LIGHT_ORANGE
                ws_out.column_dimensions[get_column_letter(calc_col_start + i)].width = 14
            
            current_row += 3
            
            # Process Data Rows
            row_in = 4
            while True:
                name_cell = ws_in.cell(row=row_in, column=1)
                raw_name = name_cell.value
                if not raw_name:
                    break
                
                clean_n = clean_name(raw_name)
                if clean_n not in monthly_stats:
                    monthly_stats[clean_n] = {'overwork': 0, 'overtime': 0, 'sundays': 0, 'days_worked': 0}
                
                c_name = ws_out.cell(row=current_row, column=1)
                c_name.value = clean_n
                c_name.font = Font(bold=True)
                c_name.border = BORDER_ALL_THIN
                
                total_hours = 0.0
                sunday_worked = False
                days_worked = 0
                
                col_ptr = 2
                
                for day_idx in range(7):
                    is_sunday = (day_idx == 6)
                    span = 1 if is_sunday else 4
                    
                    day_hours = 0.0
                    is_included = include_col_map.get(col_ptr, True)
                    
                    for k in range(span):
                        c_in = ws_in.cell(row=row_in, column=col_ptr + k)
                        c_out = ws_out.cell(row=current_row, column=col_ptr + k)
                        
                        if is_included:
                            c_out.value = c_in.value
                            if c_in.fill: c_out.fill = PatternFill(start_color=c_in.fill.start_color.index, fill_type='solid')
                            
                            val = str(c_in.value).strip() if c_in.value else ""
                            # Note: "Α" and "ΑΔΕΙΑ" are handled by parse_hours (counts as 8 hours)
                            if val and val not in ["None", "RR", "ΡΕΠΟ"]:
                                 h = parse_hours(val, clean_n)
                                 if h > 0: day_hours += h
                        else:
                            c_out.value = ""
                            c_out.fill = PatternFill(start_color="EEEEEE", fill_type="solid")
                        
                        c_out.border = BORDER_ALL_THIN
                        c_out.alignment = Alignment(horizontal='center', vertical='center')
                        c_out.font = Font(bold=True)
                    
                    total_hours += day_hours
                    if is_sunday and day_hours > 0:
                        sunday_worked = True
                    
                    if day_hours > 0:
                        days_worked += 1
                    
                    col_ptr += span
                
                # Calculate Dynamic Threshold
                # If week is "cut" (incomplete - less than 7 days included in target month), use days_worked × 8
                # Otherwise, use standard thresholds (40 hours for full-time, 20 for ΗΛΙΑΣ ΚΑΨΑΛΗΣ)
                
                # Count how many days are included in this week (for the target month)
                days_included_in_week = 0
                col_ptr_check = 2
                for day_idx in range(7):
                    is_sunday = (day_idx == 6)
                    span = 1 if is_sunday else 4
                    # Check if at least one column of this day is included
                    if any(include_col_map.get(col_ptr_check + k, True) for k in range(span)):
                        days_included_in_week += 1
                    col_ptr_check += span
                
                # If week is incomplete (cut week at start/end of month), use days_worked × 8
                if days_included_in_week < 7:
                    # Cut week: threshold = days_worked × 8
                    if clean_n.upper() == "ΗΛΙΑΣ ΚΑΨΑΛΗΣ":
                        # For ΗΛΙΑΣ ΚΑΨΑΛΗΣ, half-time: days_worked × 4
                        weekly_threshold = days_worked * 4
                    else:
                        weekly_threshold = days_worked * 8
                else:
                    # Full week: use standard thresholds
                    if clean_n.upper() == "ΗΛΙΑΣ ΚΑΨΑΛΗΣ":
                        weekly_threshold = 20
                    else:
                        weekly_threshold = 40
                
                overwork = 0
                overtime = 0
                
                if total_hours > weekly_threshold:
                    remainder = total_hours - weekly_threshold
                    overwork = min(remainder, 5)
                    if remainder > 5:
                        overtime = remainder - 5
                
                monthly_stats[clean_n]['overwork'] += overwork
                monthly_stats[clean_n]['overtime'] += overtime
                monthly_stats[clean_n]['days_worked'] += days_worked
                if sunday_worked:
                    monthly_stats[clean_n]['sundays'] += 1
                
                # Write Calculated Columns
                c_days = ws_out.cell(row=current_row, column=calc_col_start)
                c_days.value = days_worked
                c_days.alignment = Alignment(horizontal='center')
                c_days.border = BORDER_ALL_THIN
                c_days.font = Font(bold=True)
                
                c_total = ws_out.cell(row=current_row, column=calc_col_start + 1)
                c_total.value = total_hours
                c_total.alignment = Alignment(horizontal='center')
                c_total.border = BORDER_ALL_THIN
                c_total.font = Font(bold=True)
                if total_hours > 40: c_total.fill = FILL_ORANGE
                
                c_overwork = ws_out.cell(row=current_row, column=calc_col_start + 2)
                c_overwork.value = overwork
                c_overwork.alignment = Alignment(horizontal='center')
                c_overwork.border = BORDER_ALL_THIN
                c_overwork.font = Font(bold=True)
                if overwork > 0: c_overwork.fill = FILL_ORANGE
                
                c_overtime = ws_out.cell(row=current_row, column=calc_col_start + 3)
                c_overtime.value = overtime
                c_overtime.alignment = Alignment(horizontal='center')
                c_overtime.border = BORDER_ALL_THIN
                c_overtime.font = Font(bold=True)
                if overtime > 0: c_overtime.fill = FILL_LIGHT_ORANGE
                
                row_in += 1
                current_row += 1
            
            current_row += 2
            
        finally:
            os.unlink(tmp_path)
    
    # Generate Monthly Summary Table
    summary_headers = ["ΟΝΟΜΑΤΕΠΩΝΥΜΟ", "ΗΜΕΡΕΣ ΕΡΓΑΣΙΑΣ", "ΥΠΕΡΕΡΓΑΣΙΑ (h)", "ΥΠΕΡΩΡΙΕΣ(h)", "ΚΥΡΙΑΚΕΣ"]
    for i, header in enumerate(summary_headers):
        c = ws_out.cell(row=current_row, column=1 + i)
        c.value = header
        c.font = Font(bold=True)
        c.border = BORDER_THICK
        c.alignment = Alignment(horizontal='center')
        if i > 0: c.fill = FILL_HEADER_GREY
    
    current_row += 1
    
    for name, stats in monthly_stats.items():
        c = ws_out.cell(row=current_row, column=1)
        c.value = name
        c.border = BORDER_ALL_THIN
        c.fill = PatternFill(start_color="E7E6E6", fill_type="solid")
        c.font = Font(bold=True)
        
        c = ws_out.cell(row=current_row, column=2)
        c.value = stats['days_worked']
        c.alignment = Alignment(horizontal='center')
        c.border = BORDER_ALL_THIN
        c.font = Font(bold=True)
        
        c = ws_out.cell(row=current_row, column=3)
        c.value = stats['overwork']
        c.alignment = Alignment(horizontal='center')
        c.border = BORDER_ALL_THIN
        c.font = Font(bold=True)
        if stats['overwork'] > 0: c.fill = FILL_ORANGE
        
        c = ws_out.cell(row=current_row, column=4)
        c.value = stats['overtime']
        c.alignment = Alignment(horizontal='center')
        c.border = BORDER_ALL_THIN
        c.font = Font(bold=True)
        if stats['overtime'] > 0: c.fill = FILL_LIGHT_ORANGE
        
        c = ws_out.cell(row=current_row, column=5)
        c.value = stats['sundays']
        c.alignment = Alignment(horizontal='center')
        c.border = BORDER_ALL_THIN
        c.font = Font(bold=True)
        
        current_row += 1
    
    ws_out.column_dimensions['A'].width = 30
    
    # Save to bytes
    output = io.BytesIO()
    wb_out.save(output)
    output.seek(0)
    
    # Generate filename
    if target_month in month_names:
        filename = f"ΣΥΓΚΕΝΤΡΩΤΙΚΟ_ΜΙΣΘΟΔΟΣΙΑΣ_{month_names[target_month]}.xlsx"
    else:
        filename = "ΣΥΓΚΕΝΤΡΩΤΙΚΟ_ΜΙΣΘΟΔΟΣΙΑΣ.xlsx"
    
    return output, filename, monthly_stats

def get_monthly_work_days(uploaded_files, target_month):
    """
    Scans uploaded files and calculates days worked for each employee.
    Returns a dictionary: {employee_name: days_worked}
    """
    greek_months = {
        'ΙΑΝΟΥΑΡΙΟΥ': 1, 'ΦΕΒΡΟΥΑΡΙΟΥ': 2, 'ΜΑΡΤΙΟΥ': 3, 'ΑΠΡΙΛΙΟΥ': 4, 'ΜΑΙΟΥ': 5, 'ΜΑΪΟΥ': 5,
        'ΙΟΥΝΙΟΥ': 6, 'ΙΟΥΛΙΟΥ': 7, 'ΑΥΓΟΥΣΤΟΥ': 8, 'ΣΕΠΤΕΜΒΡΙΟΥ': 9, 'ΟΚΤΩΒΡΙΟΥ': 10, 'ΝΟΕΜΒΡΙΟΥ': 11, 'ΔΕΚΕΜΒΡΙΟΥ': 12
    }
    
    employee_days = {}
    
    for f in uploaded_files:
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            tmp.write(f.getvalue())
            tmp_path = tmp.name
            
        try:
            wb = openpyxl.load_workbook(tmp_path, data_only=True)
            ws = wb.active
            
            # Date filtering logic (same as payroll)
            include_col_map = {}
            col_ptr = 2
            dates_found = False
            
            for i in range(7):
                is_sunday = (i == 6)
                span = 1 if is_sunday else 4
                
                date_cell = ws.cell(row=2, column=col_ptr)
                date_val_raw = date_cell.value
                
                include_day = True
                if target_month and date_val_raw:
                    parsed_month = None
                    if hasattr(date_val_raw, 'month'):
                        parsed_month = date_val_raw.month
                        dates_found = True
                    else:
                        date_val = str(date_val_raw).strip().upper()
                        if '/' in date_val:
                            try:
                                parts = date_val.split('/')
                                if len(parts) >= 2:
                                    parsed_month = int(parts[1])
                                    dates_found = True
                            except: pass
                        elif any(m in date_val for m in greek_months.keys()):
                            for m_name, m_val in greek_months.items():
                                if m_name in date_val:
                                    parsed_month = m_val
                                    dates_found = True
                                    break
                    
                    if parsed_month and parsed_month != target_month:
                        include_day = False
                
                for k in range(span):
                    include_col_map[col_ptr + k] = include_day
                
                col_ptr += span
            
            if not dates_found and target_month:
                for col_idx in include_col_map.keys():
                    include_col_map[col_idx] = True
            
            # Scan rows for employees
            row_idx = 4
            while True:
                name_cell = ws.cell(row=row_idx, column=1)
                if not name_cell.value:
                    break
                
                clean_n = clean_name(str(name_cell.value))
                if clean_n not in employee_days:
                    employee_days[clean_n] = 0
                
                col_ptr = 2
                for day_idx in range(7):
                    is_sunday = (day_idx == 6)
                    span = 1 if is_sunday else 4
                    
                    day_hours = 0
                    is_included = include_col_map.get(col_ptr, True)
                    
                    for k in range(span):
                        if is_included:
                            c = ws.cell(row=row_idx, column=col_ptr + k)
                            val = str(c.value).strip() if c.value else ""
                            if val and val not in ["None", "RR", "ΡΕΠΟ", "ΑΝΑΡΡΩΤΙΚΗ", "ΑΔΕΙΑ"]:
                                h = parse_hours(val, clean_n)
                                if h > 0: day_hours += h
                        
                    if day_hours > 0:
                        employee_days[clean_n] += 1
                        
                    col_ptr += span
                
                row_idx += 1
                
        except Exception:
            pass
        finally:
            try: os.unlink(tmp_path)
            except: pass
            
    return employee_days

def process_cost_analysis(uploaded_files, employee_costs, target_month):
    """Process weekly schedule files and create cost analysis by location."""
    
    # Greek Month Map for Date Parsing
    greek_months = {
        'ΙΑΝΟΥΑΡΙΟΥ': 1, 'ΦΕΒΡΟΥΑΡΙΟΥ': 2, 'ΜΑΡΤΙΟΥ': 3, 'ΑΠΡΙΛΙΟΥ': 4, 'ΜΑΙΟΥ': 5, 'ΜΑΪΟΥ': 5,
        'ΙΟΥΝΙΟΥ': 6, 'ΙΟΥΛΙΟΥ': 7, 'ΑΥΓΟΥΣΤΟΥ': 8, 'ΣΕΠΤΕΜΒΡΙΟΥ': 9, 'ΟΚΤΩΒΡΙΟΥ': 10, 'ΝΟΕΜΒΡΙΟΥ': 11, 'ΔΕΚΕΜΒΡΙΟΥ': 12
    }
    
    # Sort files
    file_list = [(f.name, f) for f in uploaded_files]
    file_list.sort(key=lambda x: get_file_date_score(x[0]))
    
    # Create output workbook
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "ΚΟΣΤΟΛΟΓΗΣΗ"
    
    current_row = 1
    location_costs = {"ΡΕΝΤΗΣ": 0, "ΑΙΓΑΛΕΩ": 0, "ΠΕΙΡΑΙΑΣ": 0, "ΠΕΡΙΣΤΕΡΙ": 0}
    
    # DEBUG: Track color detections
    debug_colors = []
    
    # Process each file
    for file_name, file_obj in file_list:
        # Save uploaded file to temp location
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            tmp.write(file_obj.getvalue())
            tmp_path = tmp.name
        
        try:
            wb_in = openpyxl.load_workbook(tmp_path)
            ws_in = wb_in.active
            
            # Date filtering logic (same as payroll)
            last_data_col = 26
            include_col_map = {}
            
            col_ptr = 2
            dates_found = False
            
            for i in range(7):
                is_sunday = (i == 6)
                span = 1 if is_sunday else 4
                
                date_cell = ws_in.cell(row=2, column=col_ptr)
                date_val_raw = date_cell.value
                
                include_day = True
                if target_month and date_val_raw:
                    parsed_month = None
                    
                    if hasattr(date_val_raw, 'month'):
                        parsed_month = date_val_raw.month
                        dates_found = True
                    else:
                        date_val = str(date_val_raw).strip().upper()
                        
                        if '/' in date_val:
                            try:
                                parts = date_val.split('/')
                                if len(parts) >= 2:
                                    parsed_month = int(parts[1])
                                    dates_found = True
                            except: pass
                        
                        elif any(m in date_val for m in greek_months.keys()):
                            for m_name, m_val in greek_months.items():
                                if m_name in date_val:
                                    parsed_month = m_val
                                    dates_found = True
                                    break
                    
                    if parsed_month:
                        if parsed_month != target_month:
                            include_day = False
                
                for k in range(span):
                    include_col_map[col_ptr + k] = include_day
                
                col_ptr += span
            
            if not dates_found and target_month:
                for col_idx in include_col_map.keys():
                    include_col_map[col_idx] = True
            
            # Write Week Title
            ws_out.cell(row=current_row, column=1).value = file_name.replace("(ΕΠΙΘ).xlsx", "").replace(".xlsx", "")
            ws_out.cell(row=current_row, column=1).font = Font(bold=True, size=12)
            current_row += 1
            
            # Copy Headers (Rows 1-3)
            for r in range(1, 4):
                for c in range(1, last_data_col + 1):
                    cell_in = ws_in.cell(row=r, column=c)
                    cell_out = ws_out.cell(row=current_row + r - 1, column=c)
                    cell_out.value = cell_in.value
                    
                    if cell_in.has_style:
                        cell_out.font = Font(name=cell_in.font.name, size=cell_in.font.size, bold=True)
                        cell_out.alignment = Alignment(horizontal=cell_in.alignment.horizontal, vertical=cell_in.alignment.vertical, wrap_text=cell_in.alignment.wrap_text)
                        cell_out.border = BORDER_ALL_THIN
                        if cell_in.fill and cell_in.fill.start_color.index != '00000000':
                             cell_out.fill = PatternFill(start_color=cell_in.fill.start_color.index, fill_type='solid')
                    
                    if c in include_col_map and not include_col_map[c]:
                        cell_out.fill = PatternFill(start_color="EEEEEE", fill_type="solid")
                    
                    ws_out.column_dimensions[get_column_letter(c)].width = 16
            
            # Re-apply merges
            base_r = current_row
            col_ptr = 2
            for i in range(7):
                is_sunday = (i == 6)
                span = 1 if is_sunday else 4
                ws_out.merge_cells(start_row=base_r, start_column=col_ptr, end_row=base_r, end_column=col_ptr+span-1)
                ws_out.merge_cells(start_row=base_r+1, start_column=col_ptr, end_row=base_r+1, end_column=col_ptr+span-1)
                col_ptr += span
            
            current_row += 3
            
            # Process Data Rows - REPLACE HOURS WITH COSTS
            row_in = 4
            while True:
                name_cell = ws_in.cell(row=row_in, column=1)
                raw_name = name_cell.value
                if not raw_name:
                    break
                
                clean_n = clean_name(raw_name)
                
                # Write employee name
                c_name = ws_out.cell(row=current_row, column=1)
                c_name.value = clean_n
                c_name.font = Font(bold=True)
                c_name.border = BORDER_ALL_THIN
                
                col_ptr = 2
                
                for day_idx in range(7):
                    is_sunday = (day_idx == 6)
                    span = 1 if is_sunday else 4
                    
                    is_included = include_col_map.get(col_ptr, True)
                    
                    for k in range(span):
                        c_in = ws_in.cell(row=row_in, column=col_ptr + k)
                        c_out = ws_out.cell(row=current_row, column=col_ptr + k)
                        
                        if is_included:
                            val = str(c_in.value).strip() if c_in.value else ""
                            
                            # Check if this is work (not RR, ΡΕΠΟ, etc)
                            is_work = False
                            if val and val not in ["None", "", "RR", "ΡΕΠΟ"]:
                                if "-" in val or val.upper() in ["Α", "A", "ΑΝΑΡΡΩΤΙΚΗ", "ΑΔΕΙΑ"]:
                                    is_work = True
                            
                            # Replace with cost if this is work
                            if is_work:
                                # Get daily cost (default to 0 if not in dict)
                                daily_cost = employee_costs.get(clean_n, 0.0)
                                c_out.value = daily_cost
                                c_out.number_format = '0.00'
                                
                                # Track location cost based on COLUMN POSITION (more reliable than color)
                                if daily_cost > 0:
                                    location = None
                                    
                                    # Determine location based on column index (k)
                                    if span == 4:
                                        if k == 0: location = "ΡΕΝΤΗΣ"
                                        elif k == 1: location = "ΑΙΓΑΛΕΩ"
                                        elif k == 2: location = "ΠΕΙΡΑΙΑΣ"
                                        elif k == 3: location = "ΠΕΡΙΣΤΕΡΙ"
                                    elif span == 1:
                                        # Sunday usually has only 1 column. 
                                        # We can try to guess from header or default to RENTIS (most common)
                                        # Or check color as fallback
                                        location = "ΡΕΝΤΗΣ" # Default for Sunday
                                        
                                        # Optional: Check color just in case for Sunday
                                        if c_in.fill and hasattr(c_in.fill, 'start_color') and c_in.fill.start_color:
                                            try:
                                                color = c_in.fill.start_color.index
                                                color_clean = str(color).replace("00", "").upper()
                                                if "E2EFDA" in color_clean: location = "ΑΙΓΑΛΕΩ"
                                                elif "DDEBF7" in color_clean: location = "ΠΕΙΡΑΙΑΣ"
                                                elif "F4B084" in color_clean: location = "ΠΕΡΙΣΤΕΡΙ"
                                            except:
                                                pass
                                    
                                    if location:
                                        location_costs[location] += daily_cost
                                        
                                        # DEBUG: Track this
                                        debug_colors.append({
                                            'employee': clean_n,
                                            'cost': daily_cost,
                                            'location': location,
                                            'method': f"Column {k} (Span {span})"
                                        })
                            else:
                                # Keep original value
                                c_out.value = c_in.value
                            
                            # Copy styling
                            if c_in.fill: 
                                try:
                                    c_out.fill = PatternFill(start_color=c_in.fill.start_color.index, fill_type='solid')
                                except:
                                    pass
                        else:
                            c_out.value = ""
                            c_out.fill = PatternFill(start_color="EEEEEE", fill_type="solid")
                        
                        c_out.border = BORDER_ALL_THIN
                        c_out.alignment = Alignment(horizontal='center', vertical='center')
                        c_out.font = Font(bold=True)
                    
                    col_ptr += span
                
                row_in += 1
                current_row += 1
            
            current_row += 2
            
        finally:
            os.unlink(tmp_path)
    
    # Add ΚΟΣΤΟΣ ΑΝΑ ΚΑΤΑΣΤΗΜΑ summary
    summary_row = current_row + 1
    
    # Create fresh header cells
    for c in range(1, 5):
        cell = ws_out.cell(row=summary_row, column=c)
        cell.value = None
        cell.border = Border()
        cell.fill = PatternFill()
    
    # Merge first
    ws_out.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=4)
    
    # Then style the merged cell
    header_cell = ws_out.cell(row=summary_row, column=1)
    header_cell.value = "ΚΟΣΤΟΣ ΑΝΑ ΚΑΤΑΣΤΗΜΑ"
    header_cell.font = Font(bold=True, size=14)
    header_cell.border = Border(
        left=BORDER_THICK,
        right=BORDER_THICK,
        top=BORDER_THICK,
        bottom=BORDER_THICK
    )
    
    summary_row += 1
    
    # Data rows
    for location, cost in location_costs.items():
        # Location name
        cell_name = ws_out.cell(row=summary_row, column=1)
        cell_name.value = location
        cell_name.font = Font(bold=True)
        cell_name.border = BORDER_ALL_THIN
        cell_name.fill = FILL_HEADER_GREY
        
        # Cost value
        cell_cost = ws_out.cell(row=summary_row, column=2)
        cell_cost.value = cost
        cell_cost.number_format = '#,##0.00'
        cell_cost.font = Font(bold=True)
        cell_cost.border = BORDER_ALL_THIN
        cell_cost.alignment = Alignment(horizontal='right')
        
        summary_row += 1
    
    ws_out.column_dimensions['A'].width = 30
    
    # Save to bytes
    output = io.BytesIO()
    wb_out.save(output)
    output.seek(0)
    
    return output, location_costs, debug_colors

# === STREAMLIT UI ===
st.set_page_config(
    page_title="ThikiShop Μισθοδοσία & Κοστολόγηση", 
    page_icon="📊", 
    layout="wide",
    initial_sidebar_state="collapsed"
)

# Custom CSS for better styling
st.markdown("""
    <style>
    /* Main styling */
    .main {
        padding-top: 2rem;
    }
    
    /* Hero section */
    .hero-section {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        padding: 2rem;
        border-radius: 10px;
        color: white;
        margin-bottom: 2rem;
        box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
    }
    
    .hero-title {
        font-size: 2.5rem;
        font-weight: bold;
        margin-bottom: 0.5rem;
    }
    
    .hero-subtitle {
        font-size: 1.1rem;
        opacity: 0.9;
    }
    
    /* Step cards */
    .step-card {
        background: #f8f9fa;
        padding: 1.5rem;
        border-radius: 8px;
        border-left: 4px solid #667eea;
        margin-bottom: 1.5rem;
        box-shadow: 0 2px 4px rgba(0, 0, 0, 0.05);
    }
    
    /* Info boxes */
    .info-box {
        background: #e3f2fd;
        padding: 1rem;
        border-radius: 6px;
        border-left: 4px solid #2196f3;
        margin: 1rem 0;
    }
    
    /* Success box */
    .success-box {
        background: #e8f5e9;
        padding: 1rem;
        border-radius: 6px;
        border-left: 4px solid #4caf50;
        margin: 1rem 0;
    }
    
    /* Metrics styling */
    .metric-card {
        background: white;
        padding: 1rem;
        border-radius: 8px;
        box-shadow: 0 2px 4px rgba(0, 0, 0, 0.1);
        text-align: center;
    }
    
    /* Button styling */
    .stButton > button {
        border-radius: 6px;
        font-weight: 600;
        transition: all 0.3s;
    }
    
    .stButton > button:hover {
        transform: translateY(-2px);
        box-shadow: 0 4px 8px rgba(0, 0, 0, 0.2);
    }
    
    /* File uploader styling */
    .uploadedFile {
        background: #f0f2f6;
        border-radius: 6px;
    }
    
    /* Tab styling */
    .stTabs [data-baseweb="tab-list"] {
        gap: 8px;
    }
    
    .stTabs [data-baseweb="tab"] {
        border-radius: 8px 8px 0 0;
        padding: 10px 20px;
    }
    
    /* Footer */
    .footer {
        text-align: center;
        padding: 2rem;
        color: #666;
        margin-top: 3rem;
    }
    </style>
""", unsafe_allow_html=True)

# Hero Section
st.markdown("""
    <div class="hero-section">
        <div class="hero-title">📊 ThikiShop</div>
        <div class="hero-subtitle">Σύστημα Αυτοματοποιημένης Μισθοδοσίας & Κοστολόγησης</div>
    </div>
""", unsafe_allow_html=True)

# Create tabs
tab1, tab2 = st.tabs(["💰 Μισθοδοσία", "🏪 Κοστολόγηση Καταστημάτων"])

# === TAB 1: PAYROLL ===
with tab1:
    # Header with info
    col_header1, col_header2 = st.columns([3, 1])
    with col_header1:
        st.header("💰 Αυτόματη Δημιουργία Μισθοδοσίας")
    with col_header2:
        with st.expander("📖 Οδηγίες", expanded=False):
            st.markdown("""
            **Πώς να χρησιμοποιήσεις:**
            1. Ανέβασε τα `(ΕΠΙΘ).xlsx` αρχεία
            2. Διάλεξε τον μήνα
            3. Πάτα "Δημιουργία"
            4. Κατέβασε το αρχείο
            
            **Υπολογισμοί:**
            - Κανονικές: 40 ώρες
            - Υπερεργασία: μέχρι 5h
            - Υπερωρίες: πάνω από 45h
            """)
    
    # Step 1: File Upload
    st.markdown("### 📁 Βήμα 1: Ανέβασε τα Εβδομαδιαία Προγράμματα")
    with st.container():
        uploaded_files = st.file_uploader(
            "Επίλεξε αρχεία Excel (ΕΠΙΘ).xlsx",
            type=['xlsx'],
            accept_multiple_files=True,
            help="Μπορείς να επιλέξεις πολλά αρχεία ταυτόχρονα",
            key="payroll_upload"
        )
        
        if uploaded_files:
            st.success(f"✅ **{len(uploaded_files)}** αρχεία ανέβηκαν επιτυχώς!")
            with st.expander("📋 Προβολή αρχείων", expanded=False):
                for f in uploaded_files:
                    st.write(f"📄 {f.name}")
    
    # Step 2: Month Selection
    st.markdown("### 📅 Βήμα 2: Επιλογή Μήνα")
    month_names_display = {
        1: 'Ιανουάριος', 2: 'Φεβρουάριος', 3: 'Μάρτιος', 4: 'Απρίλιος',
        5: 'Μάιος', 6: 'Ιούνιος', 7: 'Ιούλιος', 8: 'Αύγουστος',
        9: 'Σεπτέμβριος', 10: 'Οκτώβριος', 11: 'Νοέμβριος', 12: 'Δεκέμβριος'
    }
    
    col_month1, col_month2 = st.columns([2, 3])
    with col_month1:
        selected_month = st.selectbox(
            "Επίλεξε μήνα:",
            options=list(month_names_display.keys()),
            format_func=lambda x: month_names_display[x],
            index=10,
            key="payroll_month"
        )
    with col_month2:
        st.markdown(f"<br><p style='color: #666;'>Επιλεγμένος μήνας: <strong>{month_names_display[selected_month]}</strong></p>", unsafe_allow_html=True)
    
    # Step 3: Generate Payroll
    st.markdown("### 🚀 Βήμα 3: Δημιουργία Μισθοδοσίας")
    
    col_btn1, col_btn2, col_btn3 = st.columns([1, 2, 1])
    with col_btn2:
        generate_btn = st.button("🚀 Δημιουργία Μισθοδοσίας", type="primary", use_container_width=True, key="gen_payroll")
    
    if generate_btn:
        if not uploaded_files:
            st.error("❌ **Σφάλμα:** Παρακαλώ ανέβασε τουλάχιστον ένα αρχείο!")
        else:
            with st.spinner(f"⏳ Επεξεργασία {len(uploaded_files)} αρχείων... Παρακαλώ περιμένετε..."):
                try:
                    output_file, filename, monthly_stats = process_payroll(uploaded_files, selected_month)
                    
                    st.success(f"✅ **Επιτυχία!** Το αρχείο '{filename}' δημιουργήθηκε!")
                    
                    # Store in session state for Tab 2
                    st.session_state['payroll_file'] = output_file
                    st.session_state['payroll_filename'] = filename
                    st.session_state['monthly_stats'] = monthly_stats
                    
                    # Show summary stats
                    if monthly_stats:
                        st.markdown("---")
                        st.markdown("### 📊 Συνοπτικά Στατιστικά")
                        total_employees = len(monthly_stats)
                        total_days = sum(s['days_worked'] for s in monthly_stats.values())
                        total_overwork = sum(s['overwork'] for s in monthly_stats.values())
                        total_overtime = sum(s['overtime'] for s in monthly_stats.values())
                        
                        col1, col2, col3, col4 = st.columns(4)
                        with col1:
                            st.metric("👥 Εργαζόμενοι", total_employees)
                        with col2:
                            st.metric("📅 Ημέρες Εργασίας", total_days)
                        with col3:
                            st.metric("⚡ Υπερεργασία", f"{total_overwork:.1f}h")
                        with col4:
                            st.metric("🔥 Υπερωρίες", f"{total_overtime:.1f}h")
                    
                    st.markdown("---")
                    
                    # Download Button
                    st.download_button(
                        label="📥 Κατέβασε το Αρχείο Μισθοδοσίας",
                        data=output_file,
                        file_name=filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        type="primary"
                    )
                    
                    st.info("💡 **Συμβουλή:** Μπορείς να πας στο Tab 'Κοστολόγηση Καταστημάτων' για να υπολογίσεις το κόστος ανά κατάστημα!")
                    
                except Exception as e:
                    st.error(f"❌ **Σφάλμα:** {str(e)}")
                    with st.expander("🔍 Λεπτομέρειες Σφάλματος"):
                        st.exception(e)

# === TAB 2: COST ANALYSIS ===
with tab2:
    # Header with info
    col_header1, col_header2 = st.columns([3, 1])
    with col_header1:
        st.header("🏪 Κοστολόγηση Ανά Κατάστημα")
    with col_header2:
        with st.expander("📖 Οδηγίες", expanded=False):
            st.markdown("""
            **Πώς να χρησιμοποιήσεις:**
            1. Ανέβασε τα `(ΕΠΙΘ).xlsx` αρχεία
            2. Διάλεξε τον μήνα
            3. Συμπλήρωσε μηνιαίο κόστος
            4. Πάτα "Δημιουργία"
            5. Κατέβασε το αρχείο
            
            **Υπολογισμοί:**
            - Ημερήσιο = Μηνιαίο ÷ Ημέρες
            - Κόστος ανά κατάστημα
            """)
    
    # Step 1: File Upload
    st.markdown("### 📁 Βήμα 1: Ανέβασε τα Εβδομαδιαία Προγράμματα")
    
    # Check if we can reuse from Tab 1
    if 'monthly_stats' in st.session_state:
        st.info("💡 **Συμβουλή:** Μπορείς να χρησιμοποιήσεις τα ίδια αρχεία από το Tab Μισθοδοσία ή να ανεβάσεις νέα!")
        monthly_stats = st.session_state['monthly_stats']
    else:
        monthly_stats = None
    
    with st.container():
        cost_uploaded_files = st.file_uploader(
            "Επίλεξε αρχεία Excel (ΕΠΙΘ).xlsx",
            type=['xlsx'],
            accept_multiple_files=True,
            help="Μπορείς να επιλέξεις πολλά αρχεία ταυτόχρονα",
            key="cost_upload"
        )
        
        if cost_uploaded_files:
            st.success(f"✅ **{len(cost_uploaded_files)}** αρχεία ανέβηκαν επιτυχώς!")
            with st.expander("📋 Προβολή αρχείων", expanded=False):
                for f in cost_uploaded_files:
                    st.write(f"📄 {f.name}")
    
    # Step 2: Month Selection
    st.markdown("### 📅 Βήμα 2: Επιλογή Μήνα")
    col_month1, col_month2 = st.columns([2, 3])
    with col_month1:
        cost_selected_month = st.selectbox(
            "Επίλεξε μήνα:",
            options=list(month_names_display.keys()),
            format_func=lambda x: month_names_display[x],
            index=10,
            key="cost_month"
        )
    with col_month2:
        st.markdown(f"<br><p style='color: #666;'>Επιλεγμένος μήνας: <strong>{month_names_display[cost_selected_month]}</strong></p>", unsafe_allow_html=True)
    
    # Get employee list and work days from uploaded files
    if cost_uploaded_files:
        # Calculate work days dynamically from the uploaded files
        with st.spinner("🔄 Υπολογισμός ημερών εργασίας από τα αρχεία..."):
            current_work_days = get_monthly_work_days(cost_uploaded_files, cost_selected_month)
        
        if current_work_days:
            employee_list = sorted(list(current_work_days.keys()))
            st.success(f"✅ **{len(employee_list)}** εργαζόμενοι βρέθηκαν! Οι ημέρες εργασίας υπολογίστηκαν αυτόματα.")
        else:
            st.warning("⚠️ Δεν βρέθηκαν εργαζόμενοι στα αρχεία.")
            employee_list = []

        if employee_list:
            st.markdown("### 💰 Βήμα 3: Καταχώρηση Μηνιαίου Κόστους")
            
            st.markdown("Συμπλήρωσε το **μηνιαίο κόστος** (€) για κάθε εργαζόμενο. Το ημερήσιο κόστος υπολογίζεται αυτόματα:")
            
            employee_costs = {}
            
            # Create form for employee costs
            cols = st.columns(2)
            for idx, employee_name in enumerate(employee_list):
                col = cols[idx % 2]
                
                with col:
                    # Get days worked
                    days = current_work_days.get(employee_name, 0)
                    days_info = f" ({days} ημέρες)"
                    
                    monthly_cost = st.number_input(
                        f"{employee_name}{days_info}",
                        min_value=0.0,
                        step=0.01,
                        format="%.2f",
                        key=f"cost_{employee_name}"
                    )
                    
                    if monthly_cost > 0:
                        # Calculate daily cost
                        if days > 0:
                            daily_cost = monthly_cost / days
                            employee_costs[employee_name] = daily_cost
                            st.caption(f"→ Ημερήσιο: {daily_cost:.2f}€ (÷{days})")
                        else:
                            st.warning(f"⚠️ Δεν δούλεψε καμία μέρα (Διαίρεση με 0)!")
                            employee_costs[employee_name] = 0.0
            
            st.markdown("### 🚀 Βήμα 4: Δημιουργία Κοστολόγησης")
            
            col_btn1, col_btn2, col_btn3 = st.columns([1, 2, 1])
            with col_btn2:
                generate_cost_btn = st.button("🚀 Δημιουργία Κοστολόγησης", type="primary", use_container_width=True, key="gen_cost")
            
            if generate_cost_btn:
                if not employee_costs:
                    st.error("❌ Παρακαλώ συμπλήρωσε κόστη για τουλάχιστον έναν εργαζόμενο!")
                else:
                    with st.spinner("⏳ Υπολογισμός κόστους ανά κατάστημα... Παρακαλώ περιμένετε..."):
                        try:
                            cost_file, location_costs, debug_colors = process_cost_analysis(cost_uploaded_files, employee_costs, cost_selected_month)
                            
                            st.success("✅ **Επιτυχία!** Η κοστολόγηση δημιουργήθηκε!")
                            
                            # Show summary with metrics
                            st.markdown("---")
                            st.markdown("### 📊 Κόστος Ανά Κατάστημα")
                            
                            total_cost = sum(location_costs.values())
                            
                            if total_cost == 0:
                                st.warning("⚠️ Το συνολικό κόστος είναι 0! Ελέγξτε αν τα ονόματα ταιριάζουν ακριβώς με τα αρχεία.")
                            else:
                                # Display as metrics cards
                                cols = st.columns(4)
                                locations = ["ΡΕΝΤΗΣ", "ΑΙΓΑΛΕΩ", "ΠΕΙΡΑΙΑΣ", "ΠΕΡΙΣΤΕΡΙ"]
                                for idx, loc in enumerate(locations):
                                    cost = location_costs.get(loc, 0)
                                    percentage = (cost / total_cost * 100) if total_cost > 0 else 0
                                    with cols[idx]:
                                        st.metric(
                                            loc,
                                            f"{cost:,.2f}€",
                                            delta=f"{percentage:.1f}%"
                                        )
                                
                                st.markdown("---")
                                
                                # Summary table
                                summary_data = {
                                    "Κατάστημα": list(location_costs.keys()),
                                    "Κόστος (€)": [f"{cost:,.2f}" for cost in location_costs.values()],
                                    "% Συνολικού": [f"{(cost/total_cost*100):.1f}%" for cost in location_costs.values()]
                                }
                                
                                st.dataframe(summary_data, use_container_width=True, hide_index=True)
                                
                                # Total cost metric
                                st.metric("💰 **Συνολικό Κόστος**", f"{total_cost:,.2f}€")
                            
                            # Download button
                            month_names = {
                                1: 'ΙΑΝΟΥΑΡΙΟΣ', 2: 'ΦΕΒΡΟΥΑΡΙΟΣ', 3: 'ΜΑΡΤΙΟΣ', 4: 'ΑΠΡΙΛΙΟΣ',
                                5: 'ΜΑΙΟΣ', 6: 'ΙΟΥΝΙΟΣ', 7: 'ΙΟΥΛΙΟΣ', 8: 'ΑΥΓΟΥΣΤΟΣ',
                                9: 'ΣΕΠΤΕΜΒΡΙΟΣ', 10: 'ΟΚΤΩΒΡΙΟΣ', 11: 'ΝΟΕΜΒΡΙΟΣ', 12: 'ΔΕΚΕΜΒΡΙΟΣ'
                            }
                            filename = f"ΚΟΣΤΟΛΟΓΗΣΗ_ΚΑΤΑΣΤΗΜΑΤΑ_{month_names.get(cost_selected_month, 'OUTPUT')}.xlsx"
                            
                            st.markdown("---")
                            st.download_button(
                                label="📥 Κατέβασε την Κοστολόγηση",
                                data=cost_file,
                                file_name=filename,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True,
                                type="primary"
                            )
                            
                        except Exception as e:
                            st.error(f"❌ **Σφάλμα:** {str(e)}")
                            with st.expander("🔍 Λεπτομέρειες Σφάλματος"):
                                st.exception(e)

# Footer
st.markdown("---")
st.markdown("""
    <div class="footer">
        <p><strong>ThikiShop</strong> - Σύστημα Μισθοδοσίας & Κοστολόγησης</p>
        <p style="font-size: 0.9em; color: #999;">Powered by Streamlit | Developed with ❤️</p>
    </div>
""", unsafe_allow_html=True)
